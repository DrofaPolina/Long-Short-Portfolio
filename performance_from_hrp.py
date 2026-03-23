"""
Performance from weight files and Performance workbook.

Reads:
  - Weights: Excel with sheets long_eu, short_eu, long_us, short_us (Ticker, Weight)
    Default inputs: hrp_weights.xlsx, tsfm_stock_weights.xlsx, equal_stock_weights.xlsx
  - Prices: data/Performance_SPRING_2026.xlsx (US PORTFOLIO, EU PORTFOLIO; optional EURUSD)

Writes (default):
  - outputs/portfolio_combined/performance_from_all.xlsx
    - Sheet PERFORMANCE: HRP | TSFM | EQUAL (each: EU, US, Total)
    - Weight sheets HRP_W_*, TSFM_W_*, EQUAL_W_*

Optional (--weights): a single custom output workbook (EU/US/Total sheets).

Usage:
  python performance_from_hrp.py
  python performance_from_hrp.py --weights path/to/weights.xlsx --output path/to/out.xlsx
"""
from __future__ import annotations

import argparse
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook

import config

# Paths
WEIGHTS_DIR = config.get_output_path("hrp_weights")
DATA_PATH = config.get_data_path("Performance_SPRING_2026.xlsx")
OUT_DIR = config.get_output_path("portfolio_combined")

# Weight files merged into performance_from_all.xlsx (order → columns left to right: HRP | TSFM | EQUAL).
COMBINED_WEIGHT_FILES = [
    WEIGHTS_DIR / "hrp_weights.xlsx",
    WEIGHTS_DIR / "tsfm_stock_weights.xlsx",
    WEIGHTS_DIR / "equal_stock_weights.xlsx",
]


def _portfolio_tag(weights_path: Path) -> str:
    """Sheet prefix for combined workbook: HRP, TSFM, or EQUAL."""
    name = weights_path.name.lower()
    if "equal_stock" in name:
        return "EQUAL"
    if "tsfm" in name:
        return "TSFM"
    if "hrp" in name:
        return "HRP"
    return "HRP"

# US DATA / EU DATA layout: tickers C4, prices C5, dates B5 (0-based: row 3, row 4, col 1, col 2+)
DATA_TICKER_ROW = 3
DATA_DATE_START_ROW = 4
DATA_DATE_COL = 1
DATA_VALUE_FIRST_COL = 2

# US/EU PORTFOLIO sheet layout (formula uses this sheet): dates col A, prices B:U, row 11 = first data
# Row 10 (0-based 9) = tickers in B:U, row 11 (0-based 10) = first date and first price row
PORTFOLIO_TICKER_ROW = 9   # Excel row 10
PORTFOLIO_DATE_START_ROW = 10  # Excel row 11
PORTFOLIO_DATE_COL = 0    # Column A
def _portfolio_value_cols(ncols: int) -> list[int]:
    """B:U = 20 columns; if sheet has fewer, use 1..ncols-1 (skip col 0 = date)."""
    return list(range(1, min(21, ncols)))
# Scale so first long-leg cell matches sheet AR:11 = 2500 (weights*5000, growth = P_t/P_0)
LEG_SCALE = 2500


def load_weights(path: Path) -> dict[str, pd.DataFrame]:
    """Load long/short weights from hrp_weights.xlsx (sheets: long_eu, short_eu, long_us, short_us)."""
    out = {}
    for sheet in ["long_eu", "short_eu", "long_us", "short_us"]:
        df = pd.read_excel(path, sheet_name=sheet)
        if "Ticker" in df.columns and "Weight" in df.columns:
            out[sheet] = df[["Ticker", "Weight"]].copy()
        else:
            out[sheet] = pd.DataFrame(columns=["Ticker", "Weight"])
    return out


def load_prices_from_portfolio_sheet(path: Path, sheet: str) -> pd.DataFrame:
    """
    Load prices from the PORTFOLIO sheet where the formula lives (US PORTFOLIO / EU PORTFOLIO).
    Layout: dates in column A from row 11, prices in B:U (20 cols), tickers in row 10 B:U.
    Returns price matrix for leg_performance (growth = P_t / P_0) so values match the sheet.
    """
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    value_cols = _portfolio_value_cols(raw.shape[1])
    tickers = raw.iloc[PORTFOLIO_TICKER_ROW, value_cols].astype(str).str.strip().values
    dates = pd.to_datetime(raw.iloc[PORTFOLIO_DATE_START_ROW:, PORTFOLIO_DATE_COL], errors="coerce")
    vals = raw.iloc[PORTFOLIO_DATE_START_ROW:, value_cols].copy()
    valid = dates.notna()
    dates = dates.loc[valid]
    vals = vals.loc[valid]
    vals.columns = range(vals.shape[1])
    vals = vals.apply(pd.to_numeric, errors="coerce")
    vals.index = dates.values
    vals = vals.loc[~vals.index.duplicated(keep="first")]
    vals.columns = tickers[: vals.shape[1]]
    return vals


def load_prices_from_data_sheet(path: Path, sheet: str) -> pd.DataFrame:
    """
    Load prices from US DATA / EU DATA (tickers C4, prices C5, dates B5).
    Use for short leg because PORTFOLIO sheets only have long tickers in B:U.
    """
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    value_cols = list(range(DATA_VALUE_FIRST_COL, raw.shape[1]))
    tickers = raw.iloc[DATA_TICKER_ROW, value_cols].astype(str).str.strip().values
    dates = pd.to_datetime(raw.iloc[DATA_DATE_START_ROW:, DATA_DATE_COL], errors="coerce")
    vals = raw.iloc[DATA_DATE_START_ROW:, value_cols].copy()
    valid = dates.notna()
    dates = dates.loc[valid]
    vals = vals.loc[valid]
    vals.columns = range(vals.shape[1])
    vals = vals.apply(pd.to_numeric, errors="coerce")
    vals.index = dates.values
    vals = vals.loc[~vals.index.duplicated(keep="first")]
    vals.columns = tickers[: vals.shape[1]]
    return vals


def leg_performance(weights: pd.DataFrame, prices: pd.DataFrame) -> pd.Series:
    """
    Leg performance: SUMPRODUCT(weights*LEG_SCALE, growth) with growth = P_t / P_0 (sheet formula).
    Matches sheet: (B11:U11)/($B$11:$U$11) and weights*5000 -> first cell 2500.
    """
    tickers = weights["Ticker"].astype(str).str.strip().values
    w = weights["Weight"].values
    common = [c for c in tickers if c in prices.columns]
    if not common:
        return pd.Series(index=prices.index, dtype=float)
    P = prices[common].reindex(columns=common)
    w_aligned = np.array([w[list(tickers).index(c)] for c in common])
    p0 = P.iloc[0]
    growth = P / np.where(p0 == 0, np.nan, p0)
    growth = growth.fillna(1.0)
    w_sum = w_aligned.sum()
    if abs(w_sum) < 1e-12:
        leg_value = pd.Series(0.0, index=prices.index)
    else:
        # Normalize so first row = LEG_SCALE (long) or -LEG_SCALE (short); divide by |w_sum| to preserve sign of (growth*w).sum()
        leg_value = LEG_SCALE * (growth * w_aligned).sum(axis=1) / abs(w_sum)
    return leg_value


def build_portfolio_sheet(
    long_weights: pd.DataFrame,
    short_weights: pd.DataFrame,
    long_prices: pd.DataFrame,
    short_prices: pd.DataFrame,
) -> pd.DataFrame:
    """One region: long leg from PORTFOLIO prices (match sheet), short leg from DATA prices (has short tickers)."""
    long_leg = leg_performance(long_weights, long_prices)
    short_leg = leg_performance(short_weights, short_prices)
    short_leg = short_leg.reindex(long_leg.index).fillna(0.0)
    long_leg_change = long_leg.diff()
    short_leg_change = short_leg.diff()
    # Total = long leg level + short leg level (not sum of daily deltas)
    total = long_leg + short_leg
    out = pd.DataFrame({
        "long_leg": long_leg,
        "long_leg_change": long_leg_change,
        "short_leg": short_leg,
        "short_leg_change": short_leg_change,
        "total": total,
    })
    return out


def apply_portfolio_formulas(ws, header_row: int, start_col: int, n_rows: int) -> None:
    """
    Write Excel formulas for derived columns in one portfolio table block.
    Columns (relative to start_col = Date column):
      +1 long_leg, +2 long_leg_change, +3 short_leg, +4 short_leg_change, +5 total
    total = long_leg + short_leg (levels), not long_leg_change + short_leg_change.
    """
    if n_rows <= 0:
        return
    # First data row (diff has no previous row for change columns)
    first_data_row = header_row + 1
    c_long = start_col + 1
    c_long_chg = start_col + 2
    c_short = start_col + 3
    c_short_chg = start_col + 4
    c_total = start_col + 5

    def _total_formula(r: int) -> str:
        cl = ws.cell(row=r, column=c_long).coordinate
        cs = ws.cell(row=r, column=c_short).coordinate
        return f"={cl}+{cs}"

    ws.cell(row=first_data_row, column=c_long_chg, value="=NA()")
    ws.cell(row=first_data_row, column=c_short_chg, value="=NA()")
    ws.cell(row=first_data_row, column=c_total, value=_total_formula(first_data_row))

    # Remaining rows: changes from diffs; total = long_leg + short_leg on each row
    for r in range(first_data_row + 1, first_data_row + n_rows):
        ws.cell(row=r, column=c_long_chg, value=f"={ws.cell(row=r, column=c_long).coordinate}-{ws.cell(row=r-1, column=c_long).coordinate}")
        ws.cell(row=r, column=c_short_chg, value=f"={ws.cell(row=r, column=c_short).coordinate}-{ws.cell(row=r-1, column=c_short).coordinate}")
        ws.cell(row=r, column=c_total, value=_total_formula(r))


def load_eurusd(path: Path) -> pd.Series | None:
    """Load EURUSD from Performance workbook if sheet exists; index = date, value = rate (USD per 1 EUR)."""
    xl = pd.ExcelFile(path)
    if "EURUSD" not in xl.sheet_names:
        return None
    df = pd.read_excel(path, sheet_name="EURUSD", header=None)
    # First row is header (#NAME?, EUR=); data from row 1: col 0 = date, col 1 = rate
    df = df.iloc[1:]
    dates = pd.to_datetime(df.iloc[:, 0], errors="coerce")
    vals = pd.to_numeric(df.iloc[:, 1], errors="coerce")
    s = pd.Series(vals.values, index=dates)
    s = s.loc[s.index.notna() & s.notna()]
    return s[~s.index.duplicated(keep="first")]


def run_performance(weights_path: Path, out_path: Path) -> bool:
    """
    Build performance workbook from one weight file and write one output Excel.
    Returns True if written, False if skipped (e.g. weights file missing).
    """
    if not weights_path.exists():
        print(f"Weights not found: {weights_path}")
        return False
    if not DATA_PATH.exists():
        print(f"Data file not found: {DATA_PATH}")
        return False

    weights, eu_sheet, us_sheet, total_sheet = compute_performance_frames(weights_path)
    if weights is None:
        return False

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        eu_sheet.index.name = "Date"
        us_sheet.index.name = "Date"
        total_sheet.index.name = "Date"
        eu_sheet.to_excel(w, sheet_name="EU PORTFOLIO")
        us_sheet.to_excel(w, sheet_name="US PORTFOLIO")
        total_sheet.to_excel(w, sheet_name="Total")
    # Replace derived columns with formulas in portfolio sheets
    wb = load_workbook(out_path)
    apply_portfolio_formulas(wb["EU PORTFOLIO"], header_row=1, start_col=1, n_rows=len(eu_sheet))
    apply_portfolio_formulas(wb["US PORTFOLIO"], header_row=1, start_col=1, n_rows=len(us_sheet))
    wb.save(out_path)
    print(f"Wrote {out_path}")
    return True


def compute_performance_frames(
    weights_path: Path,
) -> tuple[dict[str, pd.DataFrame] | None, pd.DataFrame | None, pd.DataFrame | None, pd.DataFrame | None]:
    """
    Compute weights and output DataFrames for one weights file.
    Returns (weights, eu_sheet, us_sheet, total_sheet), or (None, None, None, None) if skipped.
    """
    if not weights_path.exists():
        print(f"Weights not found: {weights_path}")
        return None, None, None, None
    if not DATA_PATH.exists():
        print(f"Data file not found: {DATA_PATH}")
        return None, None, None, None

    weights = load_weights(weights_path)
    long_prices_eu = load_prices_from_portfolio_sheet(DATA_PATH, "EU PORTFOLIO")
    short_prices_eu = load_prices_from_data_sheet(DATA_PATH, "EU DATA")
    long_prices_us = load_prices_from_portfolio_sheet(DATA_PATH, "US PORTFOLIO")
    short_prices_us = load_prices_from_data_sheet(DATA_PATH, "US DATA")
    eu_sheet = build_portfolio_sheet(weights["long_eu"], weights["short_eu"], long_prices_eu, short_prices_eu)
    us_sheet = build_portfolio_sheet(weights["long_us"], weights["short_us"], long_prices_us, short_prices_us)

    eurusd = load_eurusd(DATA_PATH)
    if eurusd is not None and not eurusd.empty:
        us_total = us_sheet["total"]
        fx = us_total.index.map(lambda d: eurusd.reindex([d]).iloc[0] if d in eurusd.index else np.nan)
        fx = pd.Series(fx, index=us_total.index).ffill().bfill()
        us_total_eur = us_total / fx
    else:
        us_total_eur = us_sheet["total"]

    total_sheet = pd.DataFrame({
        "eu_total": eu_sheet["total"],
        "us_total": us_sheet["total"],
        "us_total_eur": us_total_eur,
        "total_eur": eu_sheet["total"] + us_total_eur,
    })
    return weights, eu_sheet, us_sheet, total_sheet


# Single worksheet in the combined workbook: HRP | TSFM | EQUAL (EU, US, Total per method).
COMBINED_PERF_SHEET = "PERFORMANCE"


def run_performance_combined(out_path: Path) -> bool:
    """
    Build one Excel file with HRP, TSFM, and EQUAL on the same PERFORMANCE sheet
    (three horizontal blocks: each block = EU | US | Total), plus weight sheets per method.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)

    results: list[
        tuple[str, dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame, pd.DataFrame]
    ] = []
    for weights_path in COMBINED_WEIGHT_FILES:
        tag = _portfolio_tag(weights_path)
        weights, eu_sheet, us_sheet, total_sheet = compute_performance_frames(weights_path)
        if weights is None:
            continue
        eu_sheet.index.name = "Date"
        us_sheet.index.name = "Date"
        total_sheet.index.name = "Date"
        results.append((tag, weights, eu_sheet, us_sheet, total_sheet))

    if not results:
        return False

    gap = 2
    section_gap = 3  # blank columns between HRP / TSFM / EQUAL blocks
    # Row layout: model title, section label row, then table headers + data (same for all blocks).
    model_row = 0
    section_label_row = 1
    table_header_row = 2  # 0-based; Excel row 3 = column headers of EU/US/Total tables
    # 1-based Excel row of portfolio table headers (Date, long_leg, ...):
    excel_header_row_1based = table_header_row + 1

    eu0, us0, tot0 = results[0][2], results[0][3], results[0][4]
    eu_width = eu0.shape[1] + 1
    us_width = us0.shape[1] + 1
    total_width = tot0.shape[1] + 1
    block_width = eu_width + gap + us_width + gap + total_width

    wrote_any = False
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        formula_blocks: list[tuple[str, int, int, int]] = []
        current_col = 0

        for tag, weights, eu_sheet, us_sheet, total_sheet in results:
            wrote_any = True
            start_eu = current_col
            start_us = start_eu + eu_width + gap
            start_total = start_us + us_width + gap

            pd.DataFrame([[tag]]).to_excel(
                w,
                sheet_name=COMBINED_PERF_SHEET,
                index=False,
                header=False,
                startrow=model_row,
                startcol=start_eu,
            )
            pd.DataFrame({"Section": ["EU PORTFOLIO"]}).to_excel(
                w,
                sheet_name=COMBINED_PERF_SHEET,
                index=False,
                header=False,
                startrow=section_label_row,
                startcol=start_eu,
            )
            eu_sheet.to_excel(
                w, sheet_name=COMBINED_PERF_SHEET, startrow=table_header_row, startcol=start_eu
            )
            formula_blocks.append(
                (COMBINED_PERF_SHEET, excel_header_row_1based, start_eu + 1, len(eu_sheet))
            )

            pd.DataFrame({"Section": ["US PORTFOLIO"]}).to_excel(
                w,
                sheet_name=COMBINED_PERF_SHEET,
                index=False,
                header=False,
                startrow=section_label_row,
                startcol=start_us,
            )
            us_sheet.to_excel(
                w, sheet_name=COMBINED_PERF_SHEET, startrow=table_header_row, startcol=start_us
            )
            formula_blocks.append(
                (COMBINED_PERF_SHEET, excel_header_row_1based, start_us + 1, len(us_sheet))
            )

            pd.DataFrame({"Section": ["Total"]}).to_excel(
                w,
                sheet_name=COMBINED_PERF_SHEET,
                index=False,
                header=False,
                startrow=section_label_row,
                startcol=start_total,
            )
            total_sheet.to_excel(
                w, sheet_name=COMBINED_PERF_SHEET, startrow=table_header_row, startcol=start_total
            )
            formula_blocks.append(
                (COMBINED_PERF_SHEET, excel_header_row_1based, start_total + 1, len(total_sheet))
            )

            weights["long_eu"].to_excel(w, sheet_name=f"{tag}_W_long_eu", index=False)
            weights["short_eu"].to_excel(w, sheet_name=f"{tag}_W_short_eu", index=False)
            weights["long_us"].to_excel(w, sheet_name=f"{tag}_W_long_us", index=False)
            weights["short_us"].to_excel(w, sheet_name=f"{tag}_W_short_us", index=False)

            current_col += block_width + section_gap

    wb = load_workbook(out_path)
    for sheet_name, header_row, start_col, n_rows in formula_blocks:
        apply_portfolio_formulas(wb[sheet_name], header_row=header_row, start_col=start_col, n_rows=n_rows)
    wb.save(out_path)
    if wrote_any:
        print(f"Wrote combined workbook: {out_path} (sheet {COMBINED_PERF_SHEET}: HRP | TSFM | EQUAL)")
    return wrote_any


def main():
    parser = argparse.ArgumentParser(description="Build performance from weight files.")
    parser.add_argument("--weights", type=Path, help="Single weights file (run only this one)")
    parser.add_argument("--output", type=Path, help="Output Excel path (use with --weights)")
    args = parser.parse_args()

    if args.weights is not None:
        out_path = args.output if args.output is not None else OUT_DIR / "performance_from_weights.xlsx"
        run_performance(Path(args.weights), Path(out_path))
        return

    combined_out = OUT_DIR / "performance_from_all.xlsx"
    run_performance_combined(combined_out)


if __name__ == "__main__":
    main()
